'''
@Desc:   Excel工具类
@Author: Dysin
@Date:   2025/11/22
'''

import os
import pandas as pd
from PIL import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

class Excel:
    def __init__(
            self,
            path_target=None,
            name_target=None,
            bool_new=True,
            path_source=None,
            name_source=None
    ):
        self.path = path_target
        self.xlsx_name = name_target
        self.path_source = path_source
        self.name_source = name_source
        if bool_new:
            self.workbook, self.sheet = self.new_sheet()
        else:
            self.workbook, self.sheet = self.load_sheet(
                path_source,
                name_source
            )
            
    def new_sheet(self):
        # 创建一个新的Excel工作簿
        workbook = openpyxl.Workbook()
        # 选择要操作的工作表（这里选择第一个工作表，索引从0开始）
        sheet = workbook.active
        return workbook, sheet

    def load_sheet(self, path, file_name):
        # 打开工作簿
        workbook = load_workbook(os.path.join(path, file_name + '.xlsx'))
        # 选择第一个工作表
        sheet = workbook.active
        return workbook, sheet

    def style(
            self,
            font_size=12,
            wrap_text=True,
            center=True,
            auto_column_width=True,
            max_col_width=35,
            default_col_width=35,
            row_height=None,
            special_row_heights=None,
            special_col_widths=None
        ):
        '''
        通用 Excel 样式设置函数
        参数说明：
        ----------
        self.sheet : 工作表对象（openpyxl workself.sheet）
        font_size : 字体大小（默认 12）
        wrap_text : 是否自动换行（默认 True）
        center : 是否水平 + 垂直居中（默认 True）
        auto_column_width : 是否根据内容自动计算列宽（默认 True）
        max_col_width : 自动列宽的上限，防止内容过长拉伸表格（默认 35）
        default_col_width : 工作表默认列宽（默认 35）
        row_height : 设置所有行的统一行高（默认 None = 不设置）
        special_row_heights : 用于定义特殊行的行高，如 {2:100, 3:80}
        special_col_widths : 用于定义特殊列的列宽，如 {'F':20}
        '''
        # ----------------------------------------------------
        # 1. 设置单元格对齐方式、字体格式
        # ----------------------------------------------------
        # 创建一个 Alignment 对象，用于设置居中与自动换行
        alignment = Alignment(
            wrapText=wrap_text,
            horizontal='center' if center else None,
            vertical='center' if center else None
        )
        # 遍历所有单元格，应用对齐和字体格式
        for row in self.sheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = openpyxl.styles.Font(size=font_size)

        # 设置默认列宽
        self.sheet.sheet_format.defaultColWidth = default_col_width
        # ----------------------------------------------------
        # 2. 自动列宽：根据每列内容长度自动调整列宽
        # ----------------------------------------------------
        if auto_column_width:
            col_widths = []
            # 遍历每一列
            for col in range(1, self.sheet.max_column + 1):
                maxlen = 1  # 初始长度
                # 遍历当前列的所有单元格
                for row in range(1, self.sheet.max_row + 1):
                    value = self.sheet.cell(row=row, column=col).value
                    # 对数字需要考虑千分位格式
                    if isinstance(value, (int, float)):
                        length = len(str(format(value, ',')))
                    elif value is None:
                        length = 0
                    else:
                        # 中文长度要按 GBK 编码统计（避免宽度过小）
                        length = len(str(value).encode('gbk', 'ignore'))
                    maxlen = max(maxlen, length)
                col_widths.append(maxlen)
            # 设置列宽（加入 +2 作为内边距）
            for col in range(1, self.sheet.max_column + 1):
                letter = get_column_letter(col)
                self.sheet.column_dimensions[letter].width = min(col_widths[col - 1], max_col_width) + 2
        # ----------------------------------------------------
        # 3. 设置统一行高（可选）
        # ----------------------------------------------------
        if row_height is not None:
            for r in range(1, self.sheet.max_row + 1):
                self.sheet.row_dimensions[r].height = row_height
        # ----------------------------------------------------
        # 4. 设置特殊行高（如：第 2~max_row 行设为 100）
        # ----------------------------------------------------
        if special_row_heights:
            for r, h in special_row_heights.items():
                self.sheet.row_dimensions[r].height = h
        # ----------------------------------------------------
        # 5. 设置特殊列宽（如：F 列宽 = 20）
        # ----------------------------------------------------
        if special_col_widths:
            for col, w in special_col_widths.items():
                self.sheet.column_dimensions[col].width = w

    # Background color
    def background_color(self):
        filles = []
        colors = [
            'B8CCE4',
            'D8E4BC'
        ]
        for color in colors:
            filles.append(PatternFill('solid', fgColor=color))
        return filles

    # 交替背景色
    # value：根据该表头设置交替色
    def alternet(self, data, value):
        filles = self.background_color()
        fillesLast = filles[0]
        fillesNext = filles[1]
        for j in range(1, self.sheet.max_column + 1):
            self.sheet.cell(row=2, column=j).fill = filles[0]
        if (len(data) > 2):
            for i in range(1, len(data)):
                if data.loc[i, value] == data.loc[i - 1, value]:  # 根据地址
                    for j in range(1, self.sheet.max_column + 1):
                        self.sheet.cell(row=i + 2, column=j).fill = fillesLast
                else:
                    for j in range(1, self.sheet.max_column + 1):
                        self.sheet.cell(row=i + 2, column=j).fill = fillesNext
                    temp = fillesLast
                    fillesLast = fillesNext
                    fillesNext = temp

    def read_excel(self, path, file_name):
        xlsxIsExist = 1
        try:
            fileOrig = pd.read_excel(os.path.join(path, file_name + '.xlsx'), dtype=str)
        except:
            xlsxIsExist = 0
        if xlsxIsExist == 0:
            try:
                fileOrig = pd.read_excel(os.path.join(path, file_name + '.xls'), dtype=str)
            except:
                print(os.path.join(path, file_name + '.xlsx' or file_name + '.xls') + ' is not exist!!')
        data = pd.DataFrame(fileOrig)
        return data

    def read_csv(self, path, file_name):
        fileOrig = pd.read_csv(
            os.path.join(path, file_name + '.csv'),
            sep=',',
            header=0,
            encoding='gbk',
            # dtype=str
        )
        data = pd.DataFrame(fileOrig)
        return data

    def sort(self, path, file_name, columnValue):
        data = self.read_excel(path, file_name)
        data = data.sort_values(by=columnValue)
        print(data)
        data.to_excel(os.path.join(path, file_name + '.xlsx'), index=False)

    def insert_value(self, row_num, col_num, value):
        self.sheet.cell(
            row=row_num,
            column=col_num,
            value=value
        )

    def insert_row_values(self, row_num, values):
        for i in range(len(values)):
            self.sheet.cell(
                row=row_num,
                column=i+1,
                value=values[i]
            )

    def insert_header(self, values):
        self.insert_row_values(1, values)



    # 在Excel表中嵌入图片
    # file_image: 路径+图片名
    # row_num: 行号
    # column_num：列号
    def insert_image(
            self,
            row_num,
            col_num,
            file_image,
    ):
        img = ExcelImage(file_image)
        gap = 5000
        pos_str = AnchorMarker(col_num - 1, 3 * gap, row_num - 1, 3 * gap)
        pos_end = AnchorMarker(col_num, -gap, row_num, -gap)
        img.anchor = TwoCellAnchor("twoCell", pos_str, pos_end)
        self.sheet.add_image(img)

    def insert_image_to_cell(
            self,
            row_num,
            col_num,
            file_image,
            max_width=None,
            max_height=None
    ):
        """
        将图片插入指定单元格，保持长宽比
        :param sheet: Excel worksheet 对象
        :param row: 插入行号
        :param column: 插入列号
        :param file_image: 图片路径
        :param max_width: 单元格最大宽度（像素），默认 None
        :param max_height: 单元格最大高度（像素），默认 None
        """
        img = ExcelImage(file_image)
        # 获取原始尺寸
        orig_width, orig_height = img.width, img.height
        # 如果提供了最大宽高，按比例缩放
        if max_width or max_height:
            ratio_w = max_width / orig_width if max_width else 1.0
            ratio_h = max_height / orig_height if max_height else 1.0
            ratio = min(ratio_w, ratio_h, 1.0)  # 保持不放大
            img.width = int(orig_width * ratio)
            img.height = int(orig_height * ratio)
        # 计算单元格左上角坐标
        cell = self.sheet.cell(row=row_num, column=col_num)
        img.anchor = cell.coordinate  # 插入到单元格左上角
        self.sheet.add_image(img)
        # 可选：调整行高和列宽，使单元格与图片匹配
        if max_height:
            self.sheet.row_dimensions[row_num].height = max_height * 0.75  # openpyxl 行高单位约为像素*0.75
        if max_width:
            from openpyxl.utils import get_column_letter
            self.sheet.column_dimensions[get_column_letter(col_num)].width = max_width * 0.14  # 列宽约为像素*0.14

    # 在Excel中加入图片，而不是嵌入
    def add_image(self, file_image, img_size, row_num, column_num):
        img = Image(file_image)
        img.width, img.height = (img_size[0], img_size[1])
        addCell = self.sheet.cell(row=row_num, column=column_num)
        self.sheet.add_image(img, addCell.coordinate)

    def save(self):
        self.style()
        self.workbook.save(
            os.path.join(
                self.path,
                self.xlsx_name + '.xlsx'
            )
        )